'''
Created on Jun 16, 2017

@author: amschaef
'''





import tkinter as tk
from tkinter import filedialog
from tkinter import *
from skimage.io import imread,imsave
from numpy import *
from PIL import Image,ImageTk,ImageEnhance
import pandas as pd
from openpyxl import load_workbook

def _create_circle(self, x, y, r, pid, **kwargs):
    #tag_bind(id,'<ButtonPress-1>',onCircleClick)
    
    return self.create_oval(x-r, y-r, x+r, y+r, tags=pid, **kwargs)


tk.Canvas.create_circle = _create_circle


def getCoordinates(slice):
    coordinates=[]
    
    for index,row in slice.iterrows():
        if not isnan(row['x']):
            coordinates.append(row['x'])
            coordinates.append(row['y'])
    return coordinates
class Application(Frame):
    def openCorrectedData(self):
        self.filename=filedialog.askopenfilename()
        self.book=load_workbook(self.filename)
        data=pd.read_excel(self.filename,index_col=0,header=None,skiprows=[0],names=['particle','frame','x','y'],sheetname='Corrected XY Data')
        self.data=data
        self.linedict=data.groupby('particle').apply(lambda p: self.myCanvas.create_line(0,0,0,0,fill='yellow'))
        self.circledict=data.groupby('particle').apply(lambda p: self.myCanvas.create_circle(0,0,10,p.particle.values[0],outline='red',width=3))
        self.textdict=data.groupby('particle').apply(lambda p: self.myCanvas.create_text(0,0,anchor='ne',fill='red',text=p.particle.values[0],tags=p.particle.values[0]) )
        self.getPhoto(self)
    
        
    def openDataFile(self):
        self.filename=filedialog.askopenfilename()
        self.book=load_workbook(self.filename)
        data=pd.read_excel(self.filename,index_col=None,parse_cols="A:D",header=None,skiprows=[0],names=['particle','frame','x','y'])
        self.data=data
        self.linedict=data.groupby('particle').apply(lambda p: self.myCanvas.create_line(0,0,0,0,fill='yellow'))
        self.circledict=data.groupby('particle').apply(lambda p: self.myCanvas.create_circle(0,0,10,p.particle.values[0],outline='red',width=3))
        self.textdict=data.groupby('particle').apply(lambda p: self.myCanvas.create_text(0,0,anchor='ne',fill='red',text=p.particle.values[0],tags=p.particle.values[0]) )
        self.getPhoto(self)
        
    def onclick(self,event):
        
        item = self.myCanvas.find_closest(event.x, event.y)
        pid=self.myCanvas.itemcget(item,'tags')
        pid=pid.replace('current','')
        pid=int(pid)
        state=self.editvalue.get()
        if state==0: 
            self.data=self.data[self.data['particle']!=pid]
            self.myCanvas.delete(item)
            
        elif state==1:
            row=self.data.loc[(self.data['frame']==self.frame_number.get()) & (self.data['particle']==pid)]
            self.data=self.data.drop(row.index.values[0])
            self.myCanvas.delete(item)
        elif state==2:
            maxparticle=self.data['particle'].max()
            current_frame=self.frame_number.get()
            frame_data=self.data[self.data['frame']==current_frame]
           
            rowend=self.data[self.data['particle']==pid].index.values.max()
            rowstart=frame_data[frame_data['particle']==pid].index.values[0]
           
            self.data['particle'].loc[rowstart+1:rowend]=maxparticle+1    
        elif state==3:
            maxparticle=self.data['particle'].max()
            current_frame=self.frame_number.get()
            frame_data=self.data[self.data['frame']==current_frame]
           
            rowend=self.data[self.data['particle']==pid].index.values.max()
            rowstart=frame_data[frame_data['particle']==pid].index.values[0]
           
            self.data['particle'].loc[rowstart:rowend]=maxparticle+1 
        self.getPhoto(self)
    def getPhoto(self,x):
        
        currframe=self.frame_number.get()
        flat_frame=[y/15 for x in self.vid_stack[currframe] for y in x]
        flat_frame=Image.fromarray(asarray(flat_frame))
        mode_frame=flat_frame.convert('L') 
        brightenhance=self.brightness.get()
        contrastenhance=self.contrast.get()
        brightenhancer=ImageEnhance.Brightness(mode_frame)
        brightphoto=brightenhancer.enhance(brightenhance)
        contrastenhancer=ImageEnhance.Contrast(brightphoto)
        contrastphoto=contrastenhancer.enhance(contrastenhance)
        self.photo=ImageTk.PhotoImage(contrastphoto)
              
        self.myCanvas.create_image(self.myCanvas.winfo_reqwidth()/2,self.myCanvas.winfo_reqheight()/2,anchor='center',image=self.photo)
        
        if self.data.empty:
            #donothing
            return
        else:
            frame_data=self.data[self.data['frame']==currframe]
            for column, row in frame_data.iterrows():
                if not isnan(row['x']):
                    self.myCanvas.coords(self.circledict[row['particle']],[row['x']+10,row['y']+10,row['x']-10,row['y']-10])
                    self.myCanvas.coords(self.textdict[row['particle']],[row['x']+20,row['y']-20])
                else:
                    self.myCanvas.coords(self.circledict[row['particle']],[-100,-100,-100,-100])
                    self.myCanvas.coords(self.textdict[row['particle']],[-100,-100])
                self.circledict.apply(lambda p: self.myCanvas.tag_raise(p))
                self.textdict.apply(lambda p: self.myCanvas.tag_raise(p))
            if self.traceValue.get()==1: 
                past_pos=self.data[self.data['frame']<(currframe)+1]
                lines=past_pos.groupby('particle').apply(lambda p: getCoordinates(p))
                for ids in lines.index.values:
                    line=lines.ix[ids]
                    if len(line)>2:
                        lineobject=self.linedict[ids]
                        self.myCanvas.coords(lineobject,line)
                        self.myCanvas.tag_raise(lineobject)
                        
    
    def saveFile(self):
        writer=pd.ExcelWriter(self.filename, engine='openpyxl')
        sheetnames=self.book.get_sheet_names()
        if "Corrected XY Data" in sheetnames:
            ws=self.book.get_sheet_by_name("Corrected XY Data")
            self.book.remove_sheet(ws)
        
        writer.book=self.book
        writer.sheets=dict((ws.title, ws) for ws in self.book.worksheets)
        self.data.to_excel(writer,'Corrected XY Data')
        writer.save()
    
    def openVidFile(self):
        vidname = filedialog.askopenfilename()
        im=imread(vidname)
        self.num_frames=im.shape[0]  # find number of rames to split into
        self.vid_stack=(array_split(im,self.num_frames,axis=0)) #split stack into individual frames
        flat_frame=[y/15 for x in self.vid_stack[0] for y in x] # flatten first image for use, y/4 to bring values into range for tkinter?? 
        flat_frame=Image.fromarray(asarray(flat_frame))# turn into array, and then into PIL image object 
        self.photo=ImageTk.PhotoImage(flat_frame)
        self.myCanvas.create_image(self.myCanvas.winfo_reqwidth()/2,self.myCanvas.winfo_reqheight()/2,anchor='center',image=self.photo)
        self.frame_number=tk.IntVar()
        self.frameslider=tk.Scale(root,variable=self.frame_number,from_=0,to=self.num_frames-1,length=100,orient='horizontal',command=self.getPhoto)
        self.frameslider.grid(column=11,row=6,rowspan=3)
        self.framelabel=tk.Label(text='Frame Number')
        self.framelabel.grid(column=10,row=6)
        
    
    
    def createWidgets(self):
        self.myCanvas = tk.Canvas(root, width=512, height=512,highlightthickness=0)
        self.myCanvas.grid(column=1,row=1,columnspan=6,rowspan=6)
        self.myCanvas.bind('<Button-1>', self.onclick)
        #self.openVidFile()
        self.editvalue=tk.IntVar()
        self.radioa=Radiobutton(variable=self.editvalue,value=0,text='Delete Trace')
        self.radioa.grid(column=10,row=0)
        self.radiob=Radiobutton(variable=self.editvalue,value=1,text='Delete Position')
        self.radiob.grid(column=10,row=1)
        self.radioc=Radiobutton(variable=self.editvalue,value=2,text='Unlink From Next Frame')
        self.radioc.grid(column=10,row=2)
        self.radiod=Radiobutton(variable=self.editvalue,value=3,text="Unlink From Previous Frame")
        self.radiod.grid(column=10,row=3)
        self.traceValue=tk.IntVar()
        self.traceCheck=Checkbutton(root, text="Show Traces", variable=self.traceValue)
        self.traceCheck.grid(column=13,row=2)
        
        self.savebutton=Button(root,text="Save", command=self.saveFile)
        self.savebutton.grid(column=10,row=8)
        self.brightness=tk.IntVar()
        self.contrast=tk.IntVar()
        self.brightslider=tk.Scale(root,variable=self.brightness,from_=0,to=8,resolution=.1,length=100,orient='horizontal',command=self.getPhoto)
        self.contrastslider=tk.Scale(root,variable=self.contrast,from_=0,to=8,resolution=.1,length=100,orient='horizontal',command=self.getPhoto)
        self.brightlabel=tk.Label(text='Brightness')
        self.brightslider.set(1)
        self.contrastslider.set(1)
        self.contrastlabel=tk.Label(text='Contrast')
        self.brightlabel.grid(column=10,row=4)
        self.contrastlabel.grid(column=10,row=5)
        self.brightslider.grid(column=11,row=4,columnspan=2)
        self.contrastslider.grid(column=11,row=5,columnspan=2)
        self.menu=tk.Menu(root)
        root.config(menu=self.menu)
        self.filemenu = tk.Menu(self.menu)
        self.menu.add_cascade(label="Load Files", menu=self.filemenu)
        
        self.filemenu.add_command(label="Load Video", command=self.openVidFile)
        self.filemenu.add_command(label="Load Original Tracking Data", command=self.openDataFile)
        self.filemenu.add_command(label="Load Corrected Data",command=self.openCorrectedData)
    
    def __init__(self, master=None):
            Frame.__init__(self, master)
            self.filename=None
            self.data=pd.DataFrame()
            self.vid_stack=None
            self.photo=None
            self.num_frames=0
            self.book=None
            self.createWidgets()
            
           
root = Tk()
app = Application(master=root)
root.mainloop()

